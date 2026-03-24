"""
processors/excel_processor.py
────────────────────────────────────────────────────────────────────
XLSX / XLS / CSV → RAG chunk üretici.

Strateji:
  - Her sayfa (sheet) = kendi chunk grubu
  - Küçük sayfalar (≤ 60 satır) → 1 chunk
  - Büyük sayfalar → 50'şer satırlık chunk'lar
  - Her chunk başına sütun başlıkları tekrarlanır → RAG bağlamı korunur
  - CSV → tek sayfa gibi işlenir
"""

import os
import uuid

# Chunk başına maksimum satır
_ROWS_PER_CHUNK = 50


def parse_excel(
    file_path: str,
    original_name: str | None = None,
) -> list[dict]:
    file_basename = original_name or os.path.basename(file_path)
    ext = file_basename.rsplit(".", 1)[-1].lower() if "." in file_basename else ""
    chunks: list[dict] = []

    try:
        if ext == "csv":
            sheets = _read_csv(file_path)
        else:
            sheets = _read_xlsx(file_path)
    except Exception as e:
        return [{
            "id":   f"error-{uuid.uuid4()}",
            "text": f"[{file_basename}] Excel/CSV okunamadı: {e}",
            "metadata": {"source": file_basename, "error": str(e)}
        }]

    total_sheets = len(sheets)

    # ── Özet chunk: sayfa adları + satır sayıları ─────────────────────
    summary_lines = [
        f"  • {name}: {len(rows)} satır"
        for name, _, rows in sheets
    ]
    summary = (
        f"EXCEL DOSYASI: {file_basename}\n"
        f"SAYFA SAYISI: {total_sheets}\n\n"
        "── SAYFALAR ──\n"
        + "\n".join(summary_lines)
    )
    chunks.append({
        "id":   str(uuid.uuid4()),
        "text": summary,
        "metadata": {
            "page":        0,
            "chunk_index": 0,
            "source":      file_basename,
            "type":        "excel_summary",
            "total_pages": total_sheets,
        }
    })

    # ── Sayfa başına chunk'lar ────────────────────────────────────────
    global_chunk_idx = 1
    for sheet_idx, (sheet_name, headers, rows) in enumerate(sheets):
        header_line = " | ".join(str(h) for h in headers) if headers else ""

        # 50'şer satırlık batch'ler
        batches = [rows[i:i + _ROWS_PER_CHUNK] for i in range(0, max(len(rows), 1), _ROWS_PER_CHUNK)]
        total_batches = len(batches)

        for batch_idx, batch in enumerate(batches):
            row_lines = []
            for row in batch:
                row_lines.append(" | ".join(
                    _cell_str(v) for v in row
                ))

            text_parts = [
                f"[{file_basename} | Sayfa: {sheet_name} | Grup {batch_idx+1}/{total_batches}]"
            ]
            if header_line:
                text_parts.append(f"SÜTUNLAR: {header_line}")
            if row_lines:
                text_parts.append("VERİ:\n" + "\n".join(row_lines))
            else:
                text_parts.append("[Bu sayfada veri yok]")

            chunks.append({
                "id":   str(uuid.uuid4()),
                "text": "\n".join(text_parts),
                "metadata": {
                    "page":        sheet_idx + 1,
                    "chunk_index": global_chunk_idx,
                    "source":      file_basename,
                    "type":        "excel_sheet",
                    "sheet_name":  sheet_name,
                    "total_pages": total_sheets,
                }
            })
            global_chunk_idx += 1

    return chunks


# ── Yardımcı okuyucular ───────────────────────────────────────────────

def _cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _read_xlsx(file_path: str) -> list[tuple[str, list, list]]:
    """XLSX/XLS → [(sayfa_adı, başlıklar, satırlar)]"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        result = []
        for sheet_name in wb.sheetnames:
            ws     = wb[sheet_name]
            rows   = list(ws.iter_rows(values_only=True))
            if not rows:
                result.append((sheet_name, [], []))
                continue
            headers = [_cell_str(c) for c in rows[0]]
            data    = [list(r) for r in rows[1:] if any(c is not None for c in r)]
            result.append((sheet_name, headers, data))
        wb.close()
        return result
    except ImportError:
        # openpyxl yoksa xlrd dene
        import xlrd
        wb = xlrd.open_workbook(file_path)
        result = []
        for sheet in wb.sheets():
            if sheet.nrows == 0:
                result.append((sheet.name, [], []))
                continue
            headers = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
            data    = [
                [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(1, sheet.nrows)
            ]
            result.append((sheet.name, headers, data))
        return result


def _read_csv(file_path: str) -> list[tuple[str, list, list]]:
    """CSV → tek sayfa"""
    import csv
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader  = csv.reader(f)
        all_rows = list(reader)
    if not all_rows:
        return [("CSV", [], [])]
    headers = all_rows[0]
    data    = [r for r in all_rows[1:] if any(c.strip() for c in r)]
    sheet_name = os.path.splitext(os.path.basename(file_path))[0]
    return [(sheet_name, headers, data)]
